import openpyxl as xl
import Tkinter as tk
import os
import string
import ConfigParser
import shutil

class Application(tk.Frame):
    def __init__(self,master=None):

        tk.Frame.__init__(self,master)
        self.grid()
        self.createWidgets()


    def newVersion(self):
        if not os.path.isfile('config.xlsx'):
            workbook = xl.Workbook()
            worksheet = workbook.get_active_sheet()
            worksheet.title = 'config'
            worksheet.cell(row=0,column=0).value = 'Versions'
            worksheet.cell(row=0,column=1).value = 'Hardware'
            workbook.save(filename='config.xlsx')
        try:
            workbook = xl.load_workbook('config.xlsx')
        except Exception:
            errorDialog = Dialog(root,'Cannot load config.xlsx')
            root.wait_window(errorDialog.top)
        worksheet = workbook.get_active_sheet()
        i = 1
        while worksheet.cell(row=i,column=0).value != None:
            i += 1
        enterDialog = AddOption(root)
        root.wait_window(enterDialog.top)
        try:
            worksheet.cell(row=i,column=0).value = newinput
        except Exception:
            return
        workbook.save(filename='config.xlsx')

    def newHardware(self):
        if not os.path.isfile('config.xlsx'):
            workbook = xl.Workbook()
            worksheet = workbook.get_active_sheet()
            worksheet.title = 'config'
            worksheet.cell(row=0,column=0).value = 'Versions'
            worksheet.cell(row=0,column=1).value = 'Hardware'
            workbook.save(filename='config.xlsx')
        workbook = xl.load_workbook('config.xlsx')
        worksheet = workbook.get_active_sheet()
        i = 1
        while worksheet.cell(row=i,column=1).value != None:
            i += 1
        enterDialog = AddOption(root)
        root.wait_window(enterDialog.top)
        try:
            worksheet.cell(row=i,column=1).value = newinput
        except Exception:
            return
        workbook.save(filename='config.xlsx')

    def export(self):
        try:
            shutil.copyfile('Keep Out\\Master.xlsx','Assets.xlsx')
        except Exception:
            errorDialog = Dialog(root,'Error: Cannot copy.')
            root.wait_window(errorDialog.top)

    def about(self):
        ''

    def createWidgets(self):
        # Dropdown menu ##########################
        self.mb = tk.Menubutton(self,text='Menu',relief='raised')
        self.mb.grid(sticky=tk.W,columnspan=3)
        self.dropdown = tk.Menu(self.mb)
        self.mb['menu'] = self.dropdown
        self.dropdown.add_command(label='Versions',command=self.newVersion)
        self.dropdown.add_command(label='Hardware',command=self.newHardware)
        self.dropdown.add_command(label='Export',command=self.export)
        self.dropdown.add_command(label='About',command=self.about)
        
        
        # Inputs #################################
        self.inputText = ['#','A/B', 'Version', 'Hardware']
        self.inputs = [tk.StringVar() for i in range(4)]
        self.inputs[0] = tk.IntVar()
        self.inputLab = [tk.Label(self,text=self.inputText[i],width=10)
                         for i in range(4)]
        #self.inputBox = [tk.Entry(self,textvariable=self.inputs[i])
        #                 for i in range(3)]
        self.itxID = tk.Entry(self,textvariable=self.inputs[0],width=10)

        for i in range(4):
            self.inputLab[i].grid(row=1,column=i)
            #self.inputBox[i].grid(row=2,column=i)
        self.itxID.grid(row=2,column=0)

        # NEW INPUTS #############################
        self.AorBstr = tk.StringVar()
        self.AorB = tk.OptionMenu(self,self.inputs[1],'A','B')
        self.AorB.grid(row=2,column=1)
        
        self.versionstr = tk.StringVar()
        workbook = xl.load_workbook('config.xlsx')
        worksheet = workbook.get_active_sheet()
        i = 1
        versionopts = []
        while worksheet.cell(row=i,column=0).value != None:
            versionopts.append(worksheet.cell(row=i,column=0).value)
            i += 1
        self.version = tk.OptionMenu(self,self.inputs[2],
                                     *(tuple(versionopts)))
        self.version.grid(row=2,column=2)
        
        self.hardwarestr = tk.StringVar()
        i = 1
        hardwareopts = []
        while worksheet.cell(row=i,column=1).value != None:
            hardwareopts.append(worksheet.cell(row=i,column=1).value)
            i += 1
        self.hardware = tk.OptionMenu(self,self.inputs[3],
                                      *(tuple(hardwareopts)))
        self.hardware.grid(row=2,column=3)
        
        # Update button ##########################
        self.updateButton = tk.Button(command=self.update,text='Update')
        self.updateButton.grid(row=3,column=0,columnspan=3,pady=10)

    def update(self,event=None):
        # Check if necessary folder exists
        if not os.path.isdir('Keep Out'):
            os.mkdir('Keep Out')

        os.chdir('Keep Out')
        
        # Check if Excel file exists
        if not os.path.isfile('Master.xlsx'):
            workbook = xl.Workbook()
            worksheet = workbook.get_active_sheet()
            worksheet.title = 'ITX'
            headers = ['ITX','A-Version','A-Hardware','B-Version','B-Hardware']
            for i in range(5):
                worksheet.cell(row=0,column=i).value = headers[i]
                
            workbook.save(filename='Master.xlsx')

        workbook = xl.load_workbook(filename = 'Master.xlsx')
        worksheet = workbook.get_sheet_by_name(name = 'ITX')

        try:
            itx = self.inputs[0].get()
        except Exception:
            os.chdir('..')
            errorDialog = Dialog(root,'Error: Please enter a valid #')
            root.wait_window(errorDialog.top)
            return
        AorB = self.inputs[1].get()
        #print AorB
        #print itx
        '''
        if itx == '':
            os.chdir('..')
            return'''
        if string.lower(AorB) == 'a':
            self.writeline(1,2,workbook,itx)                
        elif string.lower(AorB) == 'b':
            self.writeline(3,4,workbook,itx)
            
        os.chdir('..')

    # Handle main or backup cases
    def writeline(self,a,b,workbook,itx):
        AorB = self.inputs[1].get()
        worksheet = workbook.get_sheet_by_name(name = 'ITX')
        i = 1
        exists = False
        while worksheet.cell(row=i,column=0).value != None:

            # Checks if ITX entry exists
            if int(worksheet.cell(row=i,column=0).value) == itx:

                # Opens popup to confirm overwrite
                if worksheet.cell(row=i,column=a).value != None or \
                   worksheet.cell(row=i,column=b).value != None:

                    popupstring = 'Existing entry:\nITX:'+\
                            str(itx)+string.upper(AorB)+' V: '+\
                            str(worksheet.cell(row=i,column=a).value)+' H: '+\
                            str(worksheet.cell(row=i,column=b).value)+'.'+\
                            '\n\nOverwrite?\n'
                    inputDialog = Popup(root,popupstring)
                    root.wait_window(inputDialog.top)
                else:
                    global overwrite
                    overwrite = True

                #try:
                if overwrite:
                    if str(self.inputs[2].get()) != '':
                        worksheet.cell(row=i,column=a).value = \
                            str(self.inputs[2].get())
                    if str(self.inputs[3].get()) != '':
                        worksheet.cell(row=i,column=b).value = \
                            str(self.inputs[3].get())
                    popupstring = 'Successfully entered:\n\nITX:'+\
                        str(itx)+string.upper(AorB)+' V: '+\
                        str(worksheet.cell(row=i,column=a).value)+' H: '+\
                        str(worksheet.cell(row=i,column=b).value)+'.\n'
                    successDialog = Dialog(root,popupstring)
                    root.wait_window(successDialog.top)
                #except Exception:
                    #''
                exists = True
                break
            i += 1

        if not exists:
            worksheet.cell(row=i,column=0).value = \
                            str(self.inputs[0].get())
            worksheet.cell(row=i,column=a).value = \
                            str(self.inputs[2].get())
            worksheet.cell(row=i,column=b).value = \
                            str(self.inputs[3].get())
            popupstring = 'Successfully entered:\n\nITX:'+\
                            str(itx)+string.upper(AorB)+' V: '+\
                            str(worksheet.cell(row=i,column=a).value)+' H: '+\
                            str(worksheet.cell(row=i,column=b).value)+'.\n'
            successDialog = Dialog(root,popupstring)
            root.wait_window(successDialog.top)
        workbook.save(filename='Master.xlsx')
        os.chdir('..')
        workbook.save(filename='Copy.xlsx')
        os.chdir('Keep Out')

        
class Popup:
    def __init__(self, parent, popupstring):
        top = self.top = tk.Toplevel(parent)
        self.myLabel = tk.Label(top, text=popupstring)
        self.myLabel.grid(column=0,row=0,columnspan=2)

        self.yesButton = tk.Button(top, text='Yes',
                                   command=lambda : self.send(True))
        self.yesButton.grid(row=1,column=0,sticky=tk.E)
        self.noButton = tk.Button(top, text='No',
                                  command=lambda : self.send(False))
        self.noButton.grid(row=1,column=1,sticky=tk.W)

    def send(self,doit):
        global overwrite
        overwrite = doit
        self.top.destroy()
        #print overwrite

class Dialog:
    def __init__(self, parent, popupstring):
        top = self.top = tk.Toplevel(parent)
        self.myLabel = tk.Label(top, text=popupstring)
        self.myLabel.grid(column=0,row=0)

        self.closeButton = tk.Button(top, text='Close',
                                   command=self.send)
        self.closeButton.grid(row=1,column=0)

    def send(self):
        self.top.destroy()

class AddOption:
    def __init__(self, parent):
        top = self.top = tk.Toplevel(parent)
        self.inputText = tk.StringVar()
        self.inputBox = tk.Entry(top,textvariable=self.inputText)
        self.inputBox.grid(column=0,row=0)

        self.enter = tk.Button(top, text='Enter',
                                   command=self.send)
        self.enter.grid(row=0,column=1)

    def send(self):
        global newinput
        newinput = self.inputText.get()
        self.top.destroy()
    
# MAIN PROCESS
root = tk.Tk()
#root.geometry('800x600')
#root.resizable(0,0)
#root.minsize(width=550,height=450)
app = Application()
#root.bind("<Return>",app.update)
app.master.title('ITX Info')

app.mainloop()
